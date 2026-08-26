import csv
import os
import re
from pathlib import Path

import psycopg2
from openpyxl import load_workbook

from db import getconn, putconn


def _normalize(value):
    if value is None:
        return None
    return str(value).strip()


_PHONE_RE = re.compile(r"^\+?\d[\d\s\-]{7,15}$")
_NAME_RE = re.compile(r"^[A-Za-z\s\.]{2,}$")


def _looks_like_phone(value):
    if not value:
        return False
    return bool(_PHONE_RE.match(value))


def _looks_like_name(value):
    if not value:
        return False
    return bool(_NAME_RE.match(value))


def _first_non_none(row: dict, aliases):
    lowered = {str(k).lower(): k for k in row.keys() if k is not None}
    for alias in aliases:
        if alias in lowered:
            value = _normalize(row[lowered[alias]])
            if value:
                return value
    return None


def _extract_contacts_from_row(row: dict):
    contacts = []

    rsm_name = _first_non_none(row, ["rsm name", "rsm"])
    rsm_phone = _first_non_none(row, ["ph no", "rsm\nphone number", "phone", "mobile", "mobile_no", "contact", "phone_number", "mobile_number", "sm phone number"])
    rsm_location = _first_non_none(row, ["location", "city", "branch", "branch_name", "place"])
    if rsm_name or rsm_phone:
        contacts.append({
            "name": rsm_name if _looks_like_name(rsm_name) else (rsm_phone if _looks_like_name(rsm_phone) else None),
            "phone": rsm_phone if _looks_like_phone(rsm_phone) else (rsm_name if _looks_like_phone(rsm_name) else None),
            "location": rsm_location,
            "role": "RSM",
        })

    sm_name = _first_non_none(row, ["sm name", "sm"])
    sm_phone = _first_non_none(row, ["ph no_1", "rsm\nphone number", "phone", "mobile", "mobile_no", "contact", "phone_number", "mobile_number", "sm phone number"])
    sm_location = _first_non_none(row, ["location_1", "location", "city", "branch", "branch_name", "place"])
    if sm_name or sm_phone:
        contacts.append({
            "name": sm_name if _looks_like_name(sm_name) else (sm_phone if _looks_like_name(sm_phone) else None),
            "phone": sm_phone if _looks_like_phone(sm_phone) else (sm_name if _looks_like_phone(sm_name) else None),
            "location": sm_location,
            "role": "SM",
        })

    coordinator_name = _first_non_none(row, ["coordinator", "co-ordinator", "name"])
    coordinator_phone = _first_non_none(row, ["mobile", "phone", "mobile_no", "contact", "phone_number", "mobile_number", "rsm\nphone number", "sm phone number"])
    coordinator_location = _first_non_none(row, ["location_2", "location_1", "location", "city", "branch", "branch_name", "place"])
    if coordinator_name or coordinator_phone:
        contacts.append({
            "name": coordinator_name if _looks_like_name(coordinator_name) else (coordinator_phone if _looks_like_name(coordinator_phone) else None),
            "phone": coordinator_phone if _looks_like_phone(coordinator_phone) else (coordinator_name if _looks_like_phone(coordinator_name) else None),
            "location": coordinator_location,
            "role": "Coordinator",
        })

    asm_name = _first_non_none(row, ["asm/sm/dsm name", "asm", "sm", "dsm"])
    asm_phone = _first_non_none(row, ["mobile", "phone", "mobile_no", "contact", "phone_number", "mobile_number", "rsm\nphone number", "sm phone number"])
    asm_location = _first_non_none(row, ["location_2", "location_1", "location", "city", "branch", "branch_name", "place"])
    if asm_name or asm_phone:
        contacts.append({
            "name": asm_name if _looks_like_name(asm_name) else (asm_phone if _looks_like_name(asm_phone) else None),
            "phone": asm_phone if _looks_like_phone(asm_phone) else (asm_name if _looks_like_phone(asm_name) else None),
            "location": asm_location,
            "role": "ASM/SM/DSM",
        })

    generic_name = _first_non_none(row, ["name", "manager_name", "manager", "employee_name", "contact_person"])
    generic_phone = _first_non_none(row, ["mobile", "phone", "mobile_no", "contact", "phone_number", "mobile_number", "rsm\nphone number", "sm phone number"])
    generic_location = _first_non_none(row, ["location", "city", "branch", "branch_name", "place"])
    if generic_name or generic_phone:
        contacts.append({
            "name": generic_name if _looks_like_name(generic_name) else (generic_phone if _looks_like_name(generic_phone) else None),
            "phone": generic_phone if _looks_like_phone(generic_phone) else (generic_name if _looks_like_phone(generic_name) else None),
            "location": generic_location,
            "role": None,
        })

    return contacts


def _insert_managers(conn, managers):
    cursor = conn.cursor()
    query = """
        INSERT INTO bank_managers (bank_name, name, email, phone, location, role, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """
    for m in managers:
        cursor.execute(
            query,
            (
                m.get("bank_name"),
                m.get("name") or "Unknown",
                m.get("email") or f"unknown_{id(m)}@example.com",
                m.get("phone"),
                m.get("location") or "Unknown",
                m.get("role"),
                "active",
            ),
        )
    conn.commit()
    cursor.close()


def _read_csv(path: Path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def _read_excel(path: Path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(row))
    wb.close()

    header_row_idx = None
    for idx, row in enumerate(rows):
        non_none = [v for v in row if v is not None]
        if len(non_none) >= 3 and all(isinstance(v, str) for v in non_none):
            header_row_idx = idx
            break

    if header_row_idx is None:
        header_row_idx = 0

    raw_headers = rows[header_row_idx]
    headers = []
    seen = {}
    for h in raw_headers:
        if h is None:
            headers.append(None)
            continue
        key = str(h).lower()
        if key in seen:
            seen[key] += 1
            headers.append(f"{h}_{seen[key]}")
        else:
            seen[key] = 0
            headers.append(h)

    data_rows = []
    for row in rows[header_row_idx + 1:]:
        row_dict = {}
        for i, header in enumerate(headers):
            if header is not None and i < len(row):
                row_dict[header] = row[i]
        data_rows.append(row_dict)

    return data_rows


def import_managers_from_file(file_path: str, bank_name: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        rows = _read_csv(path)
    elif ext in (".xlsx", ".xls"):
        rows = _read_excel(path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

    if not rows:
        return {"imported": 0, "skipped": 0}

    managers = []
    skipped = 0
    for row in rows:
        contacts = _extract_contacts_from_row(row)
        if not contacts:
            skipped += 1
            continue
        for contact in contacts:
            contact["bank_name"] = bank_name
            managers.append(contact)

    if not managers:
        return {"imported": 0, "skipped": skipped}

    conn = getconn()
    try:
        _insert_managers(conn, managers)
    finally:
        putconn(conn)

    return {"imported": len(managers), "skipped": skipped}
